import sys
import subprocess
import pathlib
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook
from PyQt6.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QFileDialog, QLabel, QMainWindow

######## Building GUI #########################################

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.resize(300,300)
        self.setWindowTitle("Compare Excel Files")

        # Set our Main widget that will hold everything
        MainWidget = QWidget()

######## Initialize Variables #########################################
        self.FileBtn1 = QPushButton(text="File 1")
        self.FileBtn2 = QPushButton(text="File 2")
        self.FileLabel1 = QLabel(text="File Name:")
        self.FileLabel2 = QLabel(text="File Name:")
        self.GenerateBtn = QPushButton(text="Generate")
        self.GenerateFileLabel = QLabel(text="Generated File Name:")

        
######## Create layout ################################################
        layout = QVBoxLayout()
        # Add File Button 1 to our layout 
        layout.addWidget(self.FileBtn1)
        # Add File Label 1 to our layout 
        layout.addWidget(self.FileLabel1)
        # Add File Button 2 to our layout 
        layout.addWidget(self.FileBtn2)
        # Add File Label 2 to our layout 
        layout.addWidget(self.FileLabel2)
        # Add File Generate Btn to our layout 
        layout.addWidget(self.GenerateBtn)
        # Add Generate File Label to our layout 
        layout.addWidget(self.GenerateFileLabel)

######## Apply Layout and Set Main Widget to our QMainWindow ##########

        # Apply layout to our Main Widget
        MainWidget.setLayout(layout)

        # Set our Main Widget on our QMainWindow()
        self.setCentralWidget(MainWidget)


######## Signals ################################################

        # File 1 Button Signal
        self.FileBtn1.clicked.connect(self.select_file1)
        # File 2 Button Signal
        self.FileBtn2.clicked.connect(self.select_file2)
        # Generate Button Signal
        self.GenerateBtn.clicked.connect(self.generate)
        

######## Slots ################################################

    def select_file1(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)
        if file_dialog.exec():
            self.file1_path = file_dialog.selectedFiles()[0]
            file1_name = self.file1_path.rfind('/')
            file1_name_adjusted = self.file1_path[file1_name+1:]
            self.FileLabel1.setText(f"File 1: {file1_name_adjusted}")
            
    def select_file2(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)
        if file_dialog.exec():
            self.file2_path = file_dialog.selectedFiles()[0]
            file2_name = self.file2_path.rfind('/')
            file2_name_adjusted = self.file2_path[file2_name+1:]
            self.FileLabel2.setText(f"File 2: {file2_name_adjusted}")

    def generate(self):
        if self.file1_path and self.file2_path:
            ExcelCompareMethod(self.file1_path, self.file2_path)
            # script = "Practice.py"  # Change this to the actual script name
            # cmd = f"python {script} {self.file1_path} {self.file2_path}"
            # subprocess.run(cmd, shell=True)
            mypath = pathlib.Path(r"C:\Users\tmonty\OneDrive - globalpowercomponents\Documents\VS Code\EPICOR - MAIN PROJECTS\Task 61\prac_adjusted_differences_file.xlsx")
        if mypath:
            self.GenerateFileLabel.setText("Generated File Name: adjusted_differences_file.xlsx")
        

############################################ Pandas Excel Comparison Method ################################################

############################################ Comparing Excel Files ############################################
def ExcelCompareMethod(file1, file2):

    # Specify Input Files
    file1_path = file1
    file2_path = file2

    # Create and set DataFrame variables
    # Using header= parameter, omit portions not used in the comparison 
    df1 = pd.read_excel(file1_path, header=4, index_col="MFG Part #", skipfooter=253)
    df2 = pd.read_excel(file2_path, header=4, index_col="MFG Part #", skipfooter=253)

    # Since DataFrame1 and DataFrame2 have different Indexes\Primary Keys 
    # we need to reindex DataFrame1 so it has the same Indexes\Primary Keys as DataFrame2
    df1reindex = df1.reindex(df2.index)

    # Now that both are formatted correctly we can compare
    # We use result_names= parameter to label the columns either Original or New
    dfcompare = df1reindex.compare(df2, result_names=("Original", "New"), keep_equal=True)

    # Save the differences to a new Excel file
    output_file = 'differences_file.xlsx'
    dfcompare.to_excel(output_file)  # Remove index=False

############################### Styling ############################################

    # Load the Excel workbook
    wb = openpyxl.load_workbook('differences_file.xlsx')

    # Select the worksheet where you want to apply conditional formatting
    ws = wb['Sheet1']  # Replace 'Sheet1' with the actual sheet name

    # Define the red font color for different values
    red_font = Font(color='FF0000')


    # Loop through the rows in the worksheet and change color of changed values
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=9):
        if(row[1].value != row[2].value):
            row[2].font = red_font
        if(row[3].value != row[4].value):
            row[4].font = red_font
        if(row[5].value != row[6].value):
            row[6].font = red_font
        if(row[7].value != row[8].value):
            row[8].font = red_font

    # Save the modified workbook
    wb.save('adjusted_differences_file.xlsx')
    print("\n--- File Created ---")


################################ Run GUI application loop ############################################
app = QApplication(sys.argv)
window = MainWindow()
window.show()
sys.exit(app.exec())








